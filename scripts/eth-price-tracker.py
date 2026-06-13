#!/usr/bin/env python3
"""
ETH Price Tracker - 每天获取ETH价格并更新桌面Excel表格
使用 CoinGecko 免费 API
"""
import os
import sys
import json
import datetime
import traceback

try:
    import requests
except ImportError:
    print("[ERROR] requests 库未安装，请运行: pip install requests")
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    print("[ERROR] openpyxl 库未安装，请运行: pip install openpyxl")
    sys.exit(1)

# 配置
DESKTOP_DIR = os.path.join(os.path.expanduser("~"), "Desktop")
EXCEL_FILE = os.path.join(DESKTOP_DIR, "ETH-Price-Tracker.xlsx")
# OKX API - 在中国大陆可以正常访问
OKX_TICKER_API = "https://www.okx.com/api/v5/market/ticker"

# 备用 API 列表
BACKUP_APIS = [
    "https://www.okx.com/api/v5/market/ticker",
]

HEADERS = ["日期", "时间", "ETH价格 (USD)", "24h涨跌幅 (%)"]


def fetch_eth_price_okx():
    """从 OKX 获取 ETH 价格和24小时涨跌幅"""
    params = {"instId": "ETH-USDT"}
    print(f"[INFO] 正在从 OKX 获取 ETH 价格...")
    try:
        resp = requests.get(OKX_TICKER_API, params=params, timeout=15)
        resp.raise_for_status()
        data = resp.json()

        if data.get("code") != "0":
            print(f"[ERROR] OKX API 返回错误: {data}")
            return None, None

        ticker = data.get("data", [{}])[0]
        last_price = float(ticker.get("last", 0))
        open_24h = float(ticker.get("open24h", 0))

        if last_price <= 0:
            print(f"[ERROR] 价格数据异常: {ticker}")
            return None, None

        # 计算24小时涨跌幅
        if open_24h > 0:
            change_pct = ((last_price - open_24h) / open_24h) * 100
        else:
            change_pct = 0.0

        print(f"[INFO] ETH 价格: ${last_price:.2f} USD, 24h涨跌: {change_pct:+.2f}%")
        return last_price, round(change_pct, 2)

    except requests.exceptions.RequestException as e:
        print(f"[ERROR] OKX 网络请求失败: {e}")
        return None, None
    except (ValueError, KeyError, IndexError) as e:
        print(f"[ERROR] 解析数据失败: {e}")
        return None, None


def fetch_eth_price():
    """获取 ETH 价格，自动尝试多个数据源"""
    price, change = fetch_eth_price_okx()
    if price is not None:
        return price, change

    print("[ERROR] 所有数据源均无法获取 ETH 价格")
    return None, None


def init_excel(filepath):
    """初始化 Excel 文件，如果不存在则创建"""
    if os.path.exists(filepath):
        print(f"[INFO] Excel 文件已存在: {filepath}")
        return openpyxl.load_workbook(filepath)

    print(f"[INFO] 创建新的 Excel 文件: {filepath}")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ETH价格记录"

    # 写入表头
    for col_idx, header in enumerate(HEADERS, 1):
        ws.cell(row=1, column=col_idx, value=header)

    # 设置表头样式
    from openpyxl.styles import Font, Alignment, PatternFill
    header_font = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # 设置列宽
    ws.column_dimensions["A"].width = 14  # 日期
    ws.column_dimensions["B"].width = 10  # 时间
    ws.column_dimensions["C"].width = 18  # 价格
    ws.column_dimensions["D"].width = 16  # 涨跌幅

    wb.save(filepath)
    print(f"[INFO] Excel 文件初始化完成")
    return wb


def add_price_record(filepath, price, change):
    """添加一条价格记录"""
    if not os.path.exists(filepath):
        print(f"[ERROR] Excel 文件不存在: {filepath}")
        return False

    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    now = datetime.datetime.now()
    date_str = now.strftime("%Y-%m-%d")
    time_str = now.strftime("%H:%M:%S")

    # 找到最后一行
    last_row = ws.max_row
    new_row = last_row + 1

    # 写入数据
    from openpyxl.styles import Font, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter

    ws.cell(row=new_row, column=1, value=date_str)
    ws.cell(row=new_row, column=2, value=time_str)
    ws.cell(row=new_row, column=3, value=round(price, 2))
    ws.cell(row=new_row, column=4, value=round(change, 2))

    # 设置样式
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    data_font = Font(name="Microsoft YaHei", size=10)
    center_align = Alignment(horizontal="center", vertical="center")

    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=new_row, column=col_idx)
        cell.font = data_font
        cell.alignment = center_align
        cell.border = thin_border

    # 价格列设置货币格式
    price_cell = ws.cell(row=new_row, column=3)
    price_cell.number_format = '$#,##0.00'

    # 涨跌幅列设置百分比格式
    change_cell = ws.cell(row=new_row, column=4)
    change_cell.number_format = '0.00"%"'

    # 根据涨跌设置颜色
    if change is not None:
        if change > 0:
            change_cell.font = Font(name="Microsoft YaHei", size=10, color="00B050")  # 绿色
        elif change < 0:
            change_cell.font = Font(name="Microsoft YaHei", size=10, color="FF0000")  # 红色

    wb.save(filepath)
    print(f"[INFO] 已添加记录到第 {new_row} 行: {date_str} {time_str} | ${price:.2f} | {change:.2f}%")
    return True


def main():
    print("=" * 60)
    print("  ETH 价格追踪器")
    print(f"  运行时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    # 1. 获取 ETH 价格
    price, change = fetch_eth_price()
    if price is None:
        print("[ERROR] 无法获取 ETH 价格，退出。")
        sys.exit(1)

    # 2. 初始化 Excel 文件
    try:
        wb = init_excel(EXCEL_FILE)
        wb.close()
    except Exception as e:
        print(f"[ERROR] 初始化 Excel 文件失败: {e}")
        traceback.print_exc()
        sys.exit(1)

    # 3. 添加记录
    try:
        success = add_price_record(EXCEL_FILE, price, change)
        if success:
            print(f"[SUCCESS] Excel 文件已更新: {EXCEL_FILE}")
        else:
            print("[ERROR] 添加记录失败")
            sys.exit(1)
    except Exception as e:
        print(f"[ERROR] 添加记录失败: {e}")
        traceback.print_exc()
        sys.exit(1)

    print("=" * 60)
    print("  完成！")
    print("=" * 60)


if __name__ == "__main__":
    main()
